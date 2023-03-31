# import all libraries here

import csv
import pandas as pd
import glob  # needed to get files from our folder
from datetime import datetime
import os.path
import time
import numpy as np

from dbfread import DBF  # needed to read data from DBF files
from openpyxl import Workbook, load_workbook  # needed to work with excel files

# you have to use your 'host', 'user_name', 'password', 'port' and database where you want to import files
import mysql.connector

mydb = mysql.connector.connect(host='127.0.0.1', user='root', passwd='MariaDB8813!', port=2306, db='mydatabase_hardik')
mycursor = mydb.cursor()

# create 4 different tables for 4 different Säge
mycursor.execute(
    "CREATE TABLE IF NOT EXISTS 3_saegen_machine_protokoll_ds264_1_trial3(id INT NOT NULL AUTO_INCREMENT,Zeit DATETIME ,ABS_F DOUBLE, REL_V DOUBLE, Prozent DOUBLE,v_Fed DOUBLE, v_Wir DOUBLE, N3201 DOUBLE, N3221 DOUBLE, B4003 DOUBLE, B4007 DOUBLE, B4001 DOUBLE, B4005 DOUBLE, Y2833 DOUBLE, Y2837 DOUBLE, Y2831 DOUBLE, Y2835 DOUBLE, B3801 DOUBLE, B3803 DOUBLE, Y3931 DOUBLE, Y3934 DOUBLE, B3791 DOUBLE ,B3793 DOUBLE, B2901 DOUBLE, B3905 DOUBLE, B857 DOUBLE, B3901 DOUBLE, B3903 DOUBLE, Y2841 DOUBLE, Y2843 DOUBLE, Y2845 DOUBLE, Y2847 DOUBLE, B2811 DOUBLE, Y2851 DOUBLE ,B4011 DOUBLE, B4013 DOUBLE, Visco DOUBLE, Step DOUBLE, B3821 DOUBLE, B3824 DOUBLE, B3795 DOUBLE, B3797 DOUBLE, B3801_2 DOUBLE, B3803_2 DOUBLE, M3521 DOUBLE, Y2801 DOUBLE, B3801_3 DOUBLE, B3803_3 DOUBLE, o_Fed DOUBLE, o_Wir DOUBLE, A1601_1 DOUBLE, A1611_1 DOUBLE, A1601 DOUBLE, A1611 DOUBLE, Total_power DOUBLE, Total_idle_power DOUBLE, Total_apparent_power DOUBLE, Current_Ph1 DOUBLE, Current_Ph2 DOUBLE, Current_Ph3 DOUBLE, Total_add_SM DOUBLE, Total_remove_SM DOUBLE, PRIMARY KEY (id))")
mycursor.execute(
    "CREATE TABLE IF NOT EXISTS 3_saegen_machine_protokoll_ds264_2_trial3(id INT NOT NULL AUTO_INCREMENT,Time DATETIME, ABS_F DOUBLE, REL_V DOUBLE,Prozent DOUBLE, v_Fed DOUBLE, v_Wir DOUBLE, N5301 DOUBLE, N5321 DOUBLE, B7001 DOUBLE, B7003 DOUBLE, B7005 DOUBLE, B7007 DOUBLE, Y4041 DOUBLE, Y4043 DOUBLE, Y4045 DOUBLE, Y4047 DOUBLE, B6201 DOUBLE, B6203 DOUBLE, Y6431 DOUBLE, Y6434 DOUBLE, B6021 DOUBLE, B6023 DOUBLE, B4101 DOUBLE, B6405 DOUBLE, B4021 DOUBLE, Y4061 DOUBLE, B4123 DOUBLE, Visco DOUBLE, Step DOUBLE, B6221 DOUBLE, B6224 DOUBLE, B60255 DOUBLE, B6027 DOUBLE, B6201_1 DOUBLE, B6203_1 DOUBLE, M5721 DOUBLE, Y4001 DOUBLE, B6201_2 DOUBLE, B6203_2 DOUBLE, o_Fed DOUBLE, o_Wir DOUBLE, A2701 DOUBLE, A2705 DOUBLE, Total_power DOUBLE, Total_idle_power  DOUBLE, Total_apparent_power DOUBLE, Current_Ph1 DOUBLE, Current_Ph2 DOUBLE, Current_Ph3 DOUBLE, B4023 DOUBLE, B4025 DOUBLE, B4031 DOUBLE, B4033 DOUBLE, B4035 DOUBLE, B4103 DOUBLE, Slurry_in DOUBLE, Slurry_out DOUBLE, CounterCycle DOUBLE, PRIMARY KEY (id))")
mycursor.execute(
    "CREATE TABLE IF NOT EXISTS 3_saegen_machine_protokoll_ds265_1_trial3(id INT NOT NULL AUTO_INCREMENT,Zeit DATETIME,FeedPosRel DOUBLE,Prozent DOUBLE,FeedSpeed DOUBLE, WireSpeed DOUBLE, PowerMainDrive DOUBLE, Move_Bear1_B1771 DOUBLE,Fix_Bear1_B1773 DOUBLE,Move_Bear2_B1775 DOUBLE,Fix_Bear2_B1777 DOUBLE,Slurry_B1655 DOUBLE, Valve_State_Y1751 DOUBLE,	Cool_Water_B1657 DOUBLE,Tension_N1211 DOUBLE,Tension_N1311 DOUBLE, Slurry_Flow_M1511 DOUBLE,Slurry_Visc DOUBLE, MDTorque DOUBLE,FedTorque DOUBLE,Comp_Torque DOUBLE, Comp_Angle DOUBLE,Comp_Speed DOUBLE,SlurryTempOut_B1653 DOUBLE,Total_real_power DOUBLE, Real_power_phase1 DOUBLE, Real_power_phase2 DOUBLE, Real_power_phase3 DOUBLE,PRIMARY KEY (id))")
mycursor.execute(
    "CREATE TABLE IF NOT EXISTS 3_saegen_machine_protokoll_ds265_2_trial3(id INT NOT NULL AUTO_INCREMENT,Zeit DATETIME,FeedPosRel DOUBLE,Prozent DOUBLE,FeedSpeed DOUBLE, WireSpeed DOUBLE, PowerMainDrive DOUBLE, Move_Bear1_B1771 DOUBLE,Fix_Bear1_B1773 DOUBLE,Move_Bear2_B1775 DOUBLE,Fix_Bear2_B1777 DOUBLE,Slurry_B1655 DOUBLE, Valve_State_Y1751 DOUBLE,	Cool_Water_B1657 DOUBLE,Tension_N1211 DOUBLE,Tension_N1311 DOUBLE, Slurry_Flow_M1511 DOUBLE,Slurry_Flow_Pro DOUBLE, Slurry_Dens_Pro DOUBLE, Slurry_Temp_Pro DOUBLE, Slurry_Visc DOUBLE, MDTorque DOUBLE,FedTorque DOUBLE,Comp_Torque DOUBLE, Comp_Angle DOUBLE,Comp_Speed DOUBLE,SlurryTempOut_B1653 DOUBLE,Total_real_power DOUBLE, Real_power_phase1 DOUBLE, Real_power_phase2 DOUBLE, Real_power_phase3 DOUBLE,PRIMARY KEY (id))")

mydb.commit()
# lets find a latest file from given path and work on it
# files = glob.glob("D:\Fraunhofer CSP\How to\CSP-20210623T205859Z-001\CSP/**/protocol/*.DBF")
# for dbf_table_pth in files:
files = glob.glob("D:\Fraunhofer CSP\How to\CSP-20210623T205859Z-001\CSP/**/protocol/*.DBF")
profiles = glob.glob("D:\Fraunhofer CSP\How to\CSP-20210623T205859Z-001\CSP/**/profiles/*.DBF")
for dbf_table_pth in files:
    #dbf_table_pth = max(files, key=os.path.getctime)  # latest file
    basename = os.path.basename(dbf_table_pth)
    print(dbf_table_pth)

    # Input a dbf, output a xlsx, same name, same path, except extension
    # first we convert it to csv and then it to xlsx
    csv_fn = dbf_table_pth[:-4] + ".csv"  # Set the csv file name
    table = DBF(dbf_table_pth, encoding='iso-8859-1', load=True)  # table variable is a DBF object

    # fill the data from dbf to csv
    with open(csv_fn, 'w', newline='', encoding='utf-8') as f:  # create a csv file, fill it with dbf content
        writer = csv.writer(f)
        writer.writerow(table.field_names)  # write the column name

        for record in table:  # write the rows
            writer.writerow(list(record.values()))

    # read converted csv file
    read_file = pd.read_csv(csv_fn)
    xlsx_fn = dbf_table_pth[:-4] + ".xlsx"  # Set the csv file name

    # read converted xlsx file
    read_file.to_excel(xlsx_fn, index=None, header=True)

    # USE wb class
    wb = load_workbook(xlsx_fn)
    ws = wb.active

    # iterate rows from min to max number
    rows = ws.iter_rows(ws.min_row, ws.max_row, 2, ws.max_column)
    max_col = ws.max_column  # gives maximum columns
    max_row = ws.max_row  # gives maximum rows

    # create empty list for further use
    parameters = []  # will be used to fill our data in it
    empty_columns = []  # will be used to fill empty columns whose data are not nessesary
    table_parameters = []  # will be used to fill different column names

    # lets run a loop from all rows
    for row in rows:
        for cell in row:
            ##time.sleep(0.1)
            # we want to know that which säge/machine it is, accoridng to that we extract data from source file
            # therefore get a value of cell where machine name is stated

            while cell.value == 'Machine':
                h = cell.column
                p = cell.row
                table_name = ws.cell(p, h + 1)  # machine/säge name
                print(f'\nHey, Your data from {dbf_table_pth[72:]} of Säge {table_name.value} are being imported, Get a cup of coffee!')

                # print(type(table_name.value))
                break
            # Extract data from the row where 'Time [tt:mm:jj hh:mm:ss]' is stated
            # probably it lies from row 90 to 100
            while cell.value == 'Time [tt:mm:jj hh:mm:ss]': # Time [tt:mm:jj hh:mm:ss]
                j = cell.column  # Almost lies in column 2
                i = cell.row  # 95



                # some data that are not necessary like
                # empty columms with no column names
                # or random data with no column names
                # or with column name is '-'

                # getting a row with all column names and remove unnesessary data
                rows2 = ws.iter_rows(i, ws.max_row, j, ws.max_column)
                for roow in ws.iter_rows(i, i, j, max_col):  # row with all column names
                    for cell3 in roow:
                        if cell3.value == None or cell3.value == '-':
                            K = cell3.column
                            empty_columns.append(
                                K)  # empty column is a list which fills now with data whose column names are NONE or '-'
                        else:
                            table_parameters.append(cell3.value)
                # print(empty_columns)
                # getting data now
                # bottom of the all rows there are some unnecessary data that we dont want in database
                for rw in range(i + 1, max_row + 1):
                    y = ws.cell(rw, j)
                    # print(y.value)
                    if y.value != None:  # here y detects NONE value and tells us that we dont want these datei
                        for col in range(j, max_col + 1):
                            x = ws.cell(rw, col)
                            # would take only those data which have column name
                            # and ignore other data without column names
                            if col not in empty_columns:
                                if x.value == None:  # pass values if there are no values in columns
                                    pass
                                elif col == 2:  # converts to new date format
                                    date_input = x.value
                                    datetimeobject = datetime.strptime(date_input, '%d. %b %Y %H:%M:%S')
                                    new_format = datetimeobject.strftime('%Y-%m-%d %H:%M:%S')
                                    parameters.append(new_format)
                                elif col ==3:
                                    if ws.cell(i,j+1).value == 'Feed Pos Rel [mm]':
                                        parameters.append(float(x.value))
                                        basename = os.path.basename(dbf_table_pth)
                                        basefolder = f'{dbf_table_pth[:63]}profiles\{basename}'
                                        if basefolder not in profiles:
                                            if '-' in basename:
                                                basefolder2 = f"{basefolder[:-6]}.DBF"
                                                table2 = DBF(basefolder2, encoding='GBK')
                                                df = pd.DataFrame(iter(table2))
                                                ziel = float(abs(df['PRF_FEDPOS'].iloc[-2]))
                                                quotient = (float(x.value) * 1000) / ziel
                                                parameters.append(abs(quotient * 100))
                                            else:
                                                parameters.append(None)
                                        else:
                                            table2 = DBF(basefolder, encoding='GBK')
                                            df = pd.DataFrame(iter(table2))
                                            ziel = float(abs(df['PRF_FEDPOS'].iloc[-2]))
                                            quotient = (float(x.value) * 1000) / ziel
                                            parameters.append(abs(quotient * 100))
                                    else:
                                        parameters.append(float(x.value))
                                elif col == 4:
                                    if ws.cell(i, j + 2).value == 'REL_V [mm]':
                                        parameters.append(float(x.value))
                                        basename = os.path.basename(dbf_table_pth)
                                        basefolder = f'{dbf_table_pth[:63]}profiles\{basename}'
                                        if basefolder not in profiles:
                                            if '-' in basename:
                                                basefolder2 = f"{basefolder[:-6]}.DBF"
                                                table2 = DBF(basefolder2, encoding='GBK')
                                                df = pd.DataFrame(iter(table2))
                                                ziel = float(abs(df['PRF_FEDPOS'].iloc[-2]))
                                                quotient = (float(x.value) * 1000) / ziel
                                                parameters.append(abs(quotient * 100))
                                            else:
                                                parameters.append('not available')
                                        else:
                                            table2 = DBF(basefolder, encoding='GBK')
                                            df = pd.DataFrame(iter(table2))
                                            ziel = float(abs(df['PRF_FEDPOS'].iloc[-2]))
                                            quotient = (float(x.value) * 1000) / ziel
                                            parameters.append(abs(quotient * 100))
                                    else:
                                        parameters.append(float(x.value))
                                else:
                                    parameters.append(x.value)
                            else:
                                pass
                        # now we have all the data, we have to import them into sql database
                        # but for that we must confirm which säge/machine it is
                        if table_name.value == 'DS264MC4/750 Protocol':
                            sql = "INSERT INTO 3_saegen_machine_protokoll_ds264_1_trial3(Zeit,ABS_F,REL_V, Prozent,v_Fed,v_Wir,N3201,N3221,B4003,B4007,B4001,B4005,Y2833,Y2837,Y2831,Y2835,B3801,B3803,Y3931,Y3934,B3791,B3793,B2901,B3905,B857,B3901,B3903,Y2841,Y2843,Y2845,Y2847,B2811,Y2851,B4011,B4013,Visco,Step,B3821,B3824,B3795,B3797,B3801_2,B3803_2,M3521,Y2801,B3801_3,B3803_3,o_Fed,o_Wir,A1601_1,A1611_1,A1601,A1611,Total_power,Total_idle_power,Total_apparent_power,Current_Ph1,Current_Ph2,Current_Ph3,Total_add_SM,Total_remove_SM) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
                            mycursor.execute(sql, parameters)
                        elif table_name.value == 'DS264MC 5/ 537 Protocol':
                            sql_2 = "INSERT INTO 3_saegen_machine_protokoll_ds264_2_trial3(Time, ABS_F , REL_V , Prozent, v_Fed , v_Wir , N5301 , N5321, B7001 , B7003 , B7005, B7007 , Y4041, Y4043, Y4045 , Y4047 ,B6201 ,B6203 ,Y6431 ,Y6434 ,B6021 ,B6023 ,B4101 ,B6405 ,B4021 ,Y4061 ,B4123 ,Visco ,Step ,B6221 ,B6224 ,B60255 ,B6027 ,B6201_1,B6203_1,M5721 ,Y4001 ,B6201_2,B6203_2,o_Fed ,o_Wir ,A2701 ,A2705 ,Total_power ,Total_idle_power  ,Total_apparent_power ,Current_Ph1 ,Current_Ph2 ,Current_Ph3 ,B4023 ,B4025 ,B4031 ,B4033 ,B4035 ,B4103 ,Slurry_in ,Slurry_out , CounterCycle) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
                            mycursor.execute(sql_2, parameters)
                        elif table_name.value == 'DS265MC 4/173 Protocol':
                            sql_3 = "INSERT INTO 3_saegen_machine_protokoll_ds265_1_trial3(Zeit,FeedPosRel, Prozent,FeedSpeed, WireSpeed, PowerMainDrive, Move_Bear1_B1771,Fix_Bear1_B1773,Move_Bear2_B1775,Fix_Bear2_B1777,Slurry_B1655, Valve_State_Y1751,Cool_Water_B1657,Tension_N1211,Tension_N1311, Slurry_Flow_M1511,Slurry_Visc, MDTorque,FedTorque,Comp_Torque, Comp_Angle,Comp_Speed,SlurryTempOut_B1653, Total_real_power, Real_power_phase1 , Real_power_phase2 , Real_power_phase3) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s, %s, %s, %s, %s,%s, %s, %s, %s,%s)"
                            mycursor.execute(sql_3, parameters)
                        elif table_name.value == 'DS265MC 4/178 Protocol':
                            sql_4 = "INSERT INTO 3_saegen_machine_protokoll_ds265_2_trial3(Zeit,FeedPosRel, Prozent,FeedSpeed, WireSpeed, PowerMainDrive, Move_Bear1_B1771,Fix_Bear1_B1773,Move_Bear2_B1775,Fix_Bear2_B1777,Slurry_B1655, Valve_State_Y1751,Cool_Water_B1657,Tension_N1211,Tension_N1311, Slurry_Flow_M1511,Slurry_Flow_Pro,Slurry_Dens_Pro, Slurry_Temp_Pro,Slurry_Visc, MDTorque,FedTorque,Comp_Torque, Comp_Angle,Comp_Speed,SlurryTempOut_B1653, Total_real_power, Real_power_phase1 , Real_power_phase2 , Real_power_phase3) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s, %s, %s, %s, %s,%s, %s, %s, %s, %s, %s, %s,%s)"
                            mycursor.execute(sql_4, parameters)
                        else:
                            # print(table_name.value)
                            print('Not valid data')
                        mydb.commit()
                        parameters.clear()  # clear parameters list so that it can get new values from next row

                        # print(tuple3)
                    elif y.value == None:
                        break
                    else:
                        break
                break
            break


